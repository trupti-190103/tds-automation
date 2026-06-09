import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter


def process_contractor_tds(uploaded_file):

    # =========================
    # READ EXCEL
    # =========================

    df = pd.read_excel(uploaded_file, skiprows=5)

    # CLEAN COLUMN NAMES
    df.columns = (
        df.columns
        .str.strip()
        .str.replace('\n', ' ', regex=True)
        .str.replace('  ', ' ', regex=True)
    )

    # REMOVE EMPTY ROWS
    df.dropna(how='all', inplace=True)

    # REMOVE GRAND TOTAL
    df = df[df['Particulars'].astype(str) != 'Grand Total']

    # =========================
    # EXCLUDED NON-TAXABLE COLUMNS
    # =========================

    excluded_columns = [
        'Date',
        'Particulars',
        'Voucher Type',
        'Voucher No.',
        'PAN No.',
        'Buyer/Supplier',
        'Shipping No.',
        'Shipping Date',
        'Gross Total',
        'Input CGST',
        'Input SGST',
        'Input UTGST',
        'Input IGST',
        'Sundry Balance W/off',
        'TDS on Contractor ( Individual)1023 Code'
    ]

    # KEEP ONLY EXISTING COLUMNS
    excluded_columns = [
        col for col in excluded_columns
        if col in df.columns
    ]

    # =========================
    # FIND POSSIBLE TAXABLE COLUMNS
    # =========================

    possible_columns = [
        col for col in df.columns
        if col not in excluded_columns
    ]

    # =========================
    # FIND TAXABLE VALUE
    # =========================

    def get_taxable_value(row):

        for col in possible_columns:

            value = row[col]

            if pd.notna(value):

                try:

                    numeric_value = float(value)

                    if numeric_value != 0:

                        return abs(numeric_value)

                except:

                    continue

        return 0

    # CREATE TAXABLE VALUE COLUMN
    df['Taxable Source'] = df.apply(
        get_taxable_value,
        axis=1
    )

    # =========================
    # CREATE WORKBOOK
    # =========================

    wb = Workbook()

    ws = wb.active

    ws.title = "Contractor TDS"




    # =========================
    # COPY ORIGINAL HEADER
    # =========================

    original_wb = pd.read_excel(
        uploaded_file,
        header=None,
        nrows=5
    )

    for row_index, row_data in original_wb.iterrows():
        for col_index, value in enumerate(row_data, start=1):

            ws.cell(
                row=row_index + 1,
                column=col_index
            ).value = value

    # BOLD HEADER
    bold_font = Font(bold=True)
    for row in range(1, 6):
        for col in range(1, 15):
            ws.cell(row, col).font = bold_font



    

    # =========================
    # TABLE HEADERS
    # =========================

    headers = [
        'Date',
        'Particulars',
        'Nature of Payment',
        'Code',
        'Old Section',
        'New Section',
        'Voucher No.',
        'PAN No.',
        'Taxable Value',
        'TDS on Contractor ( Individual)1023 Code',
        '%',
        'Cross Check',
        'Status',
        'Date of Payment'
    ]

    start_row = 6

    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(row=start_row, column=col_num)

        cell.value = header

        cell.font = bold_font

    # =========================
    # BORDER STYLE
    # =========================

    thin = Side(border_style="thin", color="000000")

    # =========================
    # DATA ROWS
    # =========================

    for index, row in df.iterrows():

        excel_row = start_row + 1 + index

        # DATE
        date_cell = ws.cell(excel_row, 1)

        date_cell.value = row['Date']

        date_cell.number_format = 'DD-MM-YYYY'

        # PARTICULARS
        ws.cell(excel_row, 2).value = row['Particulars']

        # NATURE OF PAYMENT
        ws.cell(excel_row, 3).value = (
            'Contractor is an Individual'
        )

        # CODE
        ws.cell(excel_row, 4).value = '1023'

        # OLD SECTION
        ws.cell(excel_row, 5).value = '194C'

        # NEW SECTION
        ws.cell(excel_row, 6).value = '391(1)'

        # VOUCHER NO
        ws.cell(excel_row, 7).value = row['Voucher No.']

        # PAN NO
        ws.cell(excel_row, 8).value = row['PAN No.']

        # TAXABLE VALUE
        ws.cell(excel_row, 9).value = row['Taxable Source']

        # ACTUAL TDS
        ws.cell(excel_row, 10).value = row[
            'TDS on Contractor ( Individual)1023 Code'
        ]

        # PERCENT
        ws.cell(excel_row, 11).value = 0.01

        ws.cell(excel_row, 11).number_format = '0.00%'

        # CROSS CHECK
        ws.cell(excel_row, 12).value = (
            f'=ROUND(I{excel_row}*K{excel_row},0)'
        )

        # STATUS
        ws.cell(excel_row, 13).value = (
            f'=IF(J{excel_row}=L{excel_row},"Correct","Mismatch")'
        )

        # DATE OF PAYMENT
        ws.cell(excel_row, 14).value = ''

        # APPLY BORDERS
        for col in range(1, 15):

            ws.cell(excel_row, col).border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )

    # =========================
    # GRAND TOTAL ROW
    # =========================

    total_row = start_row + len(df) + 1

    ws.cell(total_row, 2).value = 'Grand Total'

    ws.cell(total_row, 2).font = bold_font

    # TAXABLE VALUE TOTAL
    ws.cell(total_row, 9).value = (
        f'=SUM(I7:I{total_row-1})'
    )

    # TDS TOTAL
    ws.cell(total_row, 10).value = (
        f'=SUM(J7:J{total_row-1})'
    )

    # CROSS CHECK TOTAL
    ws.cell(total_row, 12).value = (
        f'=SUM(L7:L{total_row-1})'
    )

    # APPLY BORDERS
    for col in range(1, 15):

        ws.cell(total_row, col).border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

    # HEADER BORDERS
    for col in range(1, 15):

        ws.cell(start_row, col).border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

    # =========================
    # COLUMN WIDTHS
    # =========================

    widths = {
        1: 15,
        2: 35,
        3: 30,
        4: 12,
        5: 15,
        6: 15,
        7: 18,
        8: 18,
        9: 18,
        10: 35,
        11: 10,
        12: 18,
        13: 15,
        14: 18
    }

    for col, width in widths.items():

        ws.column_dimensions[
            get_column_letter(col)
        ].width = width

    return wb