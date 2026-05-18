import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter


def process_purchase_tds(uploaded_file):

    # READ EXCEL
    df = pd.read_excel(uploaded_file, skiprows=5)

    # REMOVE EMPTY ROWS
    df.dropna(how='all', inplace=True)

    # REMOVE GRAND TOTAL
    df = df[df['Particulars'].astype(str) != 'Grand Total']

    # KEEP REQUIRED COLUMNS
    df = df[
        [
            'Date',
            'Particulars',
            'Voucher No.',
            'PAN No.',
            'Raw Materials & Additives',
            'Discount A/c',
            'TDS on Purchase of any goods (1031 Code)'
        ]
    ]

    # CONVERT RAW MATERIAL TO POSITIVE
    df['Raw Materials & Additives'] = (
        df['Raw Materials & Additives'].abs()
    )

    # CREATE WORKBOOK
    wb = Workbook()

    ws = wb.active

    ws.title = "Purchase TDS"

    # =========================
    # HEADER SECTION
    # =========================

    ws.merge_cells('A1:P1')
    ws['A1'] = 'Plastipack (2025-2026)'

    ws.merge_cells('A2:P2')
    ws['A2'] = 'TDS on Purchase of any goods (1031 Code)'

    ws.merge_cells('A3:P3')
    ws['A3'] = 'Ledger Account'

    ws.merge_cells('A5:P5')
    ws['A5'] = '1-Apr-26 to 30-Apr-26'

    # BOLD FONT
    bold_font = Font(bold=True)

    for cell in ['A1', 'A2', 'A3', 'A5']:
        ws[cell].font = bold_font

    # =========================
    # TABLE HEADERS
    # =========================

    headers = [
        'Date',
        'Particulars',
        'Nature of Payment',
        'Code',
        'Old Section',
        'New Section',
        'Voucher No.',
        'PAN No.',
        'Raw Materials & Additives',
        'Discount A/c',
        'Taxable Value',
        'Tds on Purchase of Goods 194Q',
        '%',
        'Cross Check',
        'Status',
        'Date of Payment'
    ]

    start_row = 6

    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(row=start_row, column=col_num)

        cell.value = header

        cell.font = bold_font

    # BORDER STYLE
    thin = Side(border_style="thin", color="000000")

    # =========================
    # DATA ROWS
    # =========================

    for index, row in df.iterrows():

        excel_row = start_row + 1 + index

        # TAXABLE VALUE
        taxable_value = (
            row['Raw Materials & Additives']
            - row['Discount A/c']
        )

        # DATE
        date_cell = ws.cell(excel_row, 1)

        date_cell.value = row['Date']

        date_cell.number_format = 'DD-MM-YYYY'

        # OTHER VALUES
        ws.cell(excel_row, 2).value = row['Particulars']

        ws.cell(excel_row, 3).value = (
            'Any sum for purchase of any goods'
        )

        ws.cell(excel_row, 4).value = '1031'

        ws.cell(excel_row, 5).value = '194Q'

        ws.cell(excel_row, 6).value = '393(1)'

        ws.cell(excel_row, 7).value = row['Voucher No.']

        ws.cell(excel_row, 8).value = row['PAN No.']

        ws.cell(excel_row, 9).value = (
            row['Raw Materials & Additives']
        )

        ws.cell(excel_row, 10).value = row['Discount A/c']

        # TAXABLE VALUE
        ws.cell(excel_row, 11).value = taxable_value

        # ACTUAL TDS
        ws.cell(excel_row, 12).value = row[
            'TDS on Purchase of any goods (1031 Code)'
        ]

        # PERCENT
        ws.cell(excel_row, 13).value = 0.001

        ws.cell(excel_row, 13).number_format = '0.00%'

        # CROSS CHECK FORMULA
        ws.cell(excel_row, 14).value = (
            f'=ROUND(K{excel_row}*M{excel_row},0)'
        )

        # STATUS FORMULA
        ws.cell(excel_row, 15).value = (
            f'=IF(L{excel_row}=N{excel_row},"Correct","Mismatch")'
        )

        # DATE OF PAYMENT EMPTY
        ws.cell(excel_row, 16).value = ''

        # APPLY BORDERS
        for col in range(1, 17):

            ws.cell(excel_row, col).border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )

    # =========================
    # GRAND TOTAL ROW
    # =========================

    total_row = start_row + len(df) + 1

    ws.cell(total_row, 2).value = 'Grand Total'

    ws.cell(total_row, 2).font = bold_font

    # RAW MATERIAL TOTAL
    ws.cell(total_row, 9).value = (
        f'=SUM(I7:I{total_row-1})'
    )

    # DISCOUNT TOTAL
    ws.cell(total_row, 10).value = (
        f'=SUM(J7:J{total_row-1})'
    )

    # TAXABLE VALUE TOTAL
    ws.cell(total_row, 11).value = (
        f'=SUM(K7:K{total_row-1})'
    )

    # TDS TOTAL
    ws.cell(total_row, 12).value = (
        f'=SUM(L7:L{total_row-1})'
    )

    # APPLY BORDER TO TOTAL ROW
    for col in range(1, 17):

        ws.cell(total_row, col).border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

    # HEADER BORDERS
    for col in range(1, 17):

        ws.cell(start_row, col).border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

    # COLUMN WIDTHS
    widths = {
        1: 15,
        2: 35,
        3: 40,
        4: 12,
        5: 15,
        6: 15,
        7: 18,
        8: 18,
        9: 22,
        10: 15,
        11: 18,
        12: 25,
        13: 10,
        14: 18,
        15: 15,
        16: 18
    }

    for col, width in widths.items():

        ws.column_dimensions[
            get_column_letter(col)
        ].width = width

    return wb